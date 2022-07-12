import openpyxl
import os
import re
import json

TEMPLATE_FILENAME = "examsLinksTemplate.xlsx"
PREV_EXAMS_DIR = "./prevExamsNetSec"
OUTPUT_FILENAME = 'NetSecExamsLinks.xlsx'
COURSE_NUM = 236350
# When inserting a row, this is the default position where it will be inserted (from top of file)
MOCK_ROW = 16
FIRST_DATA_ROW = 17
ROW_INSERT_LOCATION = FIRST_DATA_ROW + 3
HEB_SEMESTER_TO_ENG = {
    "חורף": 'Winter',
    "אביב": 'Spring',
    "קיץ": 'Summer',
}
HEB_MOED_TO_ENG = {
    "א": 'A',
    "ב": 'B',
    "ג": 'C',
}

ENG_SEMESTER_TO_HEB = {
    'Winter': "חורף",
    'Spring': "אביב",
    'Summer': "קיץ",
}
ENG_MOED_TO_HEB = {
    'A': "א",
    'B': "ב",
    'C': "ג",
}
ENG_SEMESTER_TO_NUM = {
    'Winter': 1,
    'Spring': 2,
    'Summer': 3,
}


warning_text = ""
sheet = None
scansJson = None


class ColsMap:
    year = 0
    semester = 1
    moed = 2
    questions = 3
    done = 4
    sol = 5
    extra_link = 6


def get_scan_for_exam(year, semester, moed):
    global scansJson
    course_num_str = f'{COURSE_NUM}'
    if not scansJson:
        scansJson = {}
        # Load scans json from file "scans.json"
        # TODO: Load that file using firebase API (using tscans.cf api, copy from the website)
        with open('scans.json') as f:
            tmpJson = json.load(f)
        for k in tmpJson:
            v = tmpJson[k]
            term = v['term'][-2]
            sem = v['semester']
            scansJson[v['course']] = scansJson.get(v['course']) or {}
            scansJson[v['course']][sem] = scansJson[v['course']].get(sem) or {}
            scansJson[v['course']][sem][term] = scansJson[v['course']][sem].get(term) \
                or []
            scansJson[v['course']][sem][term].append(
                (tmpJson[k]['grade'], f"https://drive.google.com/file/d/{k}/view"))

        # Uncomment to save memory:
        # Filter just courses that we're currently working on
        # scansJson = {
        #     k: scansJson[k] for k in scansJson if k == course_num_str}

    semester_str = str((year - 1)*100+ENG_SEMESTER_TO_NUM[semester])
    matches = scansJson.get(course_num_str, {}).get(
        semester_str, {}).get(ENG_MOED_TO_HEB[moed], None)

    # for k in scansJson:
    #     if (scansJson[k]['course'] == course_num_str and
    #         scansJson[k]['semester'] == semester_str and
    #             ENG_MOED_TO_HEB[moed] == scansJson[k]['term'][-1]):
    #         matches.append(
    #             (scansJson[k]['grade'], f"{scansJson[k]['grade']}: https://drive.google.com/file/d/'{k}/view"))
    if matches:
        matches.sort(key=lambda x: x[0], reverse=True)
        return [x[1] for x in matches]


def path_to_year_semester_moed_is_solution(path):
    global warning_text
    # Replace all hebrew letters with coresponding latin
    path = path.replace('חורף', 'Winter')
    path = path.replace('אביב', 'Spring')
    path = path.replace('מועד', 'Moed')
    path = path.replace('פתרון', 'Solution')
    path = path.replace('answer', 'Solution')
    path = path.replace('בוחן אמצע', 'Midterm')
    path = path.replace('בוחן', 'Midterm')
    path = path.replace('אמצע', 'Midterm')
    path = path.replace('א', 'A')
    path = path.replace('ב', 'B')

    try:
        # Ignore the PREV_EXAMS_DIR part of the path
        path = path[path.find(PREV_EXAMS_DIR)+len(PREV_EXAMS_DIR):]
        if re.search(r'midterm', path, re.IGNORECASE) is not None or 'skip' in path:

            # Skip midterms and skips
            warning_text += f'Skipped midterm path: {path}\n'
            return None

        part = '/'

        if 'part' in path.lower():
            regex_res_obj = re.search(
                r'part.?([1-2]|[a-b])', path, re.IGNORECASE)
            if regex_res_obj:
                result = regex_res_obj.group(1)
                if result.lower() == "a":
                    part = 1
                elif result.lower() == "b":
                    part = 2
                else:
                    part = int(result)

        if 'DS_Store' in path or 'idterm' in path.lower():
            return
        if "sp" in path.lower():
            term = 'Spring'
        elif 'w' in path.lower():
            term = 'Winter'
        else:
            raise Exception("Term not found")

        # extract year in the form of [0-9]{2,4}
        year = int("20"+re.search(r'\d\d{2,4}(?!\d)(?!(-\d\d))', path).group(0)[-2:])

        # Extract moed from path (follows the work Moed ignore case)
        moed = re.search(
            r'(Moed|/|Exam|\d\d|Maman|Spring|Winter|Summer).?([ABC])', path, re.IGNORECASE).group(2).upper()
        # Check if "sol" apear in the path case insensitive
        is_sol = re.search(r'sol', path, re.IGNORECASE) is not None
        print(
            f'extracted {year=} {term=} {moed=} {is_sol=} {part=} \t:\t{path=}')
        return year, term, moed, is_sol, part
    except Exception as e:
        print(f'Failed to extract from {path=}.')
        print('Please fix the file name to the form SpringMoedA2022')
        raise e


def get_link_col(is_sol, part):
    if part == 2:
        return ColsMap.extra_link
    elif is_sol:
        return ColsMap.sol
    else:
        return ColsMap.questions


modified_cells = []


def set_link_cell(link_cell, path, moed, is_sol, part, is_http_url_sol=False):
    global modified_cells
    global warning_text

    if link_cell in modified_cells:
        warning_text += f"Tried overriding an already updated cell {link_cell.coordinate} ; {path=}\n"
        return

    cell_value = "Moed " + moed + (" Sol" if is_sol else "")
    if part in [1, 2]:
        cell_value += f" Pt{part}"

    modified_cells.append(link_cell)
    if is_http_url_sol:
        link_cell.value = "scan: " + cell_value
        link_cell.hyperlink = path
    else:
        link_cell.value = cell_value
        link_cell.hyperlink = path[2:]
    link_cell.style = 'Hyperlink'


def find_cell_and_add_link(path, year, semester, moed, is_sol, part, is_http_url_sol=False):
    sheet = get_sheet()
    for row in sheet.iter_rows(min_row=FIRST_DATA_ROW):
        #  Find the cell that matches the parameters
        if row[0].value == year and \
            row[1].value and HEB_SEMESTER_TO_ENG[row[1].value] == semester and \
                row[2].value and HEB_MOED_TO_ENG[row[2].value] == moed:
            link_col_num = get_link_col(is_sol=is_sol, part=part)
            set_link_cell(link_cell=row[link_col_num],
                          path=path,
                          moed=moed,
                          is_sol=is_sol,
                          part=part,
                          is_http_url_sol=is_http_url_sol)
            # Excel conditional formatting number > 2015
            # conditional_formula = '=IF(AND(ISNUMBER(A{row}),A{row}>2015),"",A{row})'
            break


def get_sheet():
    global sheet, workbook
    if sheet:
        return sheet
    workbook = openpyxl.load_workbook(TEMPLATE_FILENAME)
    sheet = workbook.active
    return sheet


def save_sheet():
    global sheet, workbook, TEMPLATE_FILENAME
    if sheet is None:
        raise Exception("Failed to save, sheet is not initialized")

    workbook.save(OUTPUT_FILENAME)
    TEMPLATE_FILENAME = OUTPUT_FILENAME
    sheet = None


def copy_row_style(from_row, to_row):
    for i, cell in enumerate(from_row):
        to_row[i]._style = cell._style


def fill_missing_data():
    print(f"Filling missing data in cells (missing years/semesters according to previous cells)")
    # Open excel file
    sheet = get_sheet()

    # Fill in missing data
    last_row = sheet[FIRST_DATA_ROW]
    not_found_count = 0
    for row in sheet.iter_rows(min_row=FIRST_DATA_ROW):
        if not_found_count >= 4:
            break

        if row[0].value == None:
            not_found_count += 1
            for i in range(2):
                if row[i].value == None:
                    row[i].value = last_row[i].value
                    row[i]._style = last_row[i]._style

            last_row = row
            continue

        not_found_count = 0
        last_row = row


def get_scan_from_tscans_if_solution_is_missing(file_details, details):
    if details[-2] == True:
        return

    for d in file_details:
        if d[1:] == (*details[1:4], True, '/'):
            return
    else:
        return get_scan_for_exam(*details[1:4])


def main():
    global warning_text

    fill_missing_data()

    print(f'Iterating over {PREV_EXAMS_DIR=}')
    file_details = []

    # Iterate files over the prev_exams_dir recursively
    sheet = get_sheet()

    for subdir, dirs, files in os.walk(PREV_EXAMS_DIR):
        for file in files:
            path = os.path.join(subdir, file)
            ret = path_to_year_semester_moed_is_solution(path)
            if not ret:
                warning_text += f'Skipped path: {path}\n'
                continue
            file_details.append((path, *ret))
            year, semester, moed, is_sol, part = ret
            if (moed == 'C'):
                print(f'Adding C term exam row: {path}')
                sheet.insert_rows(ROW_INSERT_LOCATION)
                newrow = sheet[f"{ROW_INSERT_LOCATION}"]
                copy_row_style(from_row=sheet[f"{MOCK_ROW}"], to_row=newrow)
                newrow[ColsMap.year].value = year
                newrow[ColsMap.moed].value = ENG_MOED_TO_HEB[moed]
                newrow[ColsMap.semester].value = ENG_SEMESTER_TO_HEB[semester]
                newrow[ColsMap.sol].value = "/"
                newrow[ColsMap.extra_link].value = "חסר"
                newrow[ColsMap.questions].value = "חסר"

    save_sheet()
    get_sheet()

    for details in file_details:
        path, year, semester, moed, is_sol, part = details
        find_cell_and_add_link(path=path,
                               year=year,
                               semester=semester,
                               moed=moed,
                               is_sol=is_sol,
                               part=part)
        if not is_sol:
            scans = get_scan_from_tscans_if_solution_is_missing(
                file_details, details)

            if scans:
                find_cell_and_add_link(path=scans[0],
                                       year=year,
                                       semester=semester,
                                       moed=moed,
                                       is_sol=True,
                                       part='/',
                                       is_http_url_sol=True)

    save_sheet()
    if warning_text:
        print("WARNING:\n" + warning_text)


if __name__ == "__main__":
    main()
